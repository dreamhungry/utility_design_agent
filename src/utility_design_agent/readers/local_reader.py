"""本地 Excel / CSV 文件读取器"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from ..models import NPCData
from .base import BaseReader


class LocalReader(BaseReader):
    """读取本地 .xlsx 或 .csv 策划文档"""

    async def read(self, *, path: str = "", **kwargs: Any) -> list[NPCData]:
        """读取本地文件并返回 NPCData 列表

        Args:
            path: 文件路径（.xlsx 或 .csv）
        """
        if not path:
            raise ValueError("必须提供 path 参数")

        filepath = Path(path)
        if not filepath.exists():
            raise FileNotFoundError(f"文件不存在: {filepath}")

        suffix = filepath.suffix.lower()
        if suffix == ".xlsx":
            rows = self._read_xlsx(filepath)
        elif suffix == ".csv":
            rows = self._read_csv(filepath)
        else:
            raise ValueError(f"不支持的文件格式: {suffix}，仅支持 .xlsx 和 .csv")

        return self._parse_rows(rows)

    # ------------------------------------------------------------------
    # 文件解析
    # ------------------------------------------------------------------

    @staticmethod
    def _read_xlsx(filepath: Path) -> list[list[Any]]:
        from openpyxl import load_workbook

        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):  # type: ignore[union-attr]
            rows.append(list(row))
        wb.close()
        return rows

    @staticmethod
    def _read_csv(filepath: Path) -> list[list[Any]]:
        rows: list[list[Any]] = []
        # 尝试不同编码
        for encoding in ("utf-8-sig", "utf-8", "gbk", "gb2312"):
            try:
                with filepath.open("r", encoding=encoding, newline="") as f:
                    reader = csv.reader(f)
                    rows = [list(row) for row in reader]
                break
            except UnicodeDecodeError:
                continue
        return rows

    # ------------------------------------------------------------------
    # 行解析
    # ------------------------------------------------------------------

    @staticmethod
    def _detect_columns(header: list[str]) -> dict[str, int]:
        mapping: dict[str, list[str]] = {
            "name": ["npc名称", "name", "npc_name", "名称", "npc"],
            "personality_tags": ["性格标签", "personality_tags", "personality", "标签"],
            "needs": ["需求", "needs", "behavior", "偏好"],
            "design_intent": ["设计意图", "design_intent", "intent", "意图"],
        }
        lower_header = [str(h).strip().lower() for h in header]
        result: dict[str, int] = {}
        for field, candidates in mapping.items():
            for c in candidates:
                if c in lower_header:
                    result[field] = lower_header.index(c)
                    break
        return result

    @classmethod
    def _parse_rows(cls, rows: list[list[Any]]) -> list[NPCData]:
        if len(rows) < 2:
            return []

        header = [str(c) for c in rows[0]]
        col_map = cls._detect_columns(header)
        if "name" not in col_map:
            raise ValueError(
                f"无法识别表头中的 NPC 名称列，当前表头: {header}。"
                f"支持的列名: npc名称, name, npc_name, 名称, npc"
            )

        npcs: list[NPCData] = []
        for row in rows[1:]:
            def _cell(field: str) -> Any:
                idx = col_map.get(field)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            name = _cell("name")
            if not name:
                continue

            def _split_tags(raw: Any) -> list[str]:
                if isinstance(raw, list):
                    return [str(t).strip() for t in raw if str(t).strip()]
                if isinstance(raw, str):
                    for sep in ["、", ",", "，", ";", "；"]:
                        if sep in raw:
                            return [t.strip() for t in raw.split(sep) if t.strip()]
                    return [raw.strip()] if raw.strip() else []
                return []

            npcs.append(
                NPCData(
                    name=str(name).strip(),
                    personality_tags=_split_tags(_cell("personality_tags")),
                    needs=_split_tags(_cell("needs")),
                    design_intent=str(_cell("design_intent") or ""),
                )
            )
        return npcs
